import asyncio
import logging
import os
from datetime import datetime, timedelta, time, date as date_type

import openpyxl
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)

# ─────────────────────────────────────────────
BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
OWNER_ID  = int(os.environ.get("OWNER_ID", "0"))
EXCEL_FILE = "appointments.xlsx"
# ─────────────────────────────────────────────

# Conversation states
(CHOOSE_LANG, WAITING_FOR_YES, ASK_NAME, ASK_AGE,
 ASK_REASON, ASK_MOBILE, ASK_DATE) = range(7)

# Language strings
STRINGS = {
    "en": {
        "welcome":        "🏥 *Welcome to Jivandeep Clinic!*\n\nWould you like to book an appointment?\nReply *YES* to proceed.",
        "choose_lang":    "🌐 Please choose your language:\n\nભાષા પસંદ કરો:",
        "ask_name":       "👤 Please enter your *Full Name*:",
        "ask_age":        "🔢 Please enter your *Age*:",
        "ask_reason":     "📋 Please enter your *Reason for visit*:",
        "ask_mobile":     "📱 Please enter your *Mobile Number*:\n(10 digits, Indian number e.g. 9876543210)",
        "ask_date":       "📅 Book for *TODAY* or a *SPECIFIC DATE*?\nReply *Today* or enter date as *DD/MM/YYYY*",
        "confirmed":      "✅ *Appointment Confirmed!*\n\n👤 *Name:* {name}\n🔢 *Age:* {age}\n📋 *Reason:* {reason}\n📱 *Mobile:* {mobile}\n📅 *Date:* {date}\n⏰ *Slot:* {slot}\n\nThank you for choosing *Jivandeep Clinic!* 🏥",
        "no_slots":       "😔 No available slots for that date. Please choose another date (DD/MM/YYYY):",
        "past_date":      "⚠️ Cannot book for a past date. Please choose today or a future date:",
        "invalid_date":   "⚠️ Invalid format. Reply *Today* or enter as *DD/MM/YYYY* (e.g. 25/12/2026):",
        "invalid_age":    "⚠️ Invalid age. Please enter a number between 1 and 120:",
        "invalid_mobile": "⚠️ Invalid mobile number. Please enter exactly 10 digits (e.g. 9876543210):",
        "invalid_name":   "⚠️ Name too short. Please enter your full name (at least 2 characters):",
        "closed_sunday":  "⚠️ Clinic is closed on Sundays after 1 PM. Please choose another date:",
        "reply_yes":      "Please reply *YES* to proceed. 😊",
        "cancelled":      "❌ Booking cancelled. Send /start to book again.",
        "no_booking":     "ℹ️ You have no active appointment.",
        "your_appt":      "📋 *Your Appointment*\n\n👤 {name}\n📅 {date}\n⏰ {slot}\n📋 {reason}",
        "cancel_confirm": "Are you sure you want to cancel your appointment on {date} at {slot}?",
        "cancel_done":    "✅ Your appointment has been cancelled successfully.",
        "reminder_1day":  "⏰ *Appointment Reminder*\n\nHello {name}! Your appointment at Jivandeep Clinic is *tomorrow* ({date}) at *{slot}*.\n\nPlease arrive 10 minutes early. 🏥",
        "reminder_1hr":   "⏰ *Appointment Reminder*\n\nHello {name}! Your appointment at Jivandeep Clinic is in *1 hour* at *{slot}*.\n\nPlease get ready! 🏥",
    },
    "gu": {
        "welcome":        "🏥 *જીવનદીપ ક્લિનિકમાં આપનું સ્વાગત છે!*\n\nએપોઇન્ટમેન્ટ બૂક કરવા માટે *YES* જવાબ આપો.",
        "choose_lang":    "🌐 Please choose your language:\n\nભાષા પસંદ કરો:",
        "ask_name":       "👤 કૃપા કરીને તમારું *પૂરું નામ* દાખલ કરો:",
        "ask_age":        "🔢 તમારી *ઉંમર* દાખલ કરો:",
        "ask_reason":     "📋 મુલાકાતનું *કારણ* દાખલ કરો:",
        "ask_mobile":     "📱 તમારો *મોબાઇલ નંબર* દાખલ કરો:\n(10 અંક, દા.ત. 9876543210)",
        "ask_date":       "📅 *આજ* અથવા *ચોક્કસ તારીખ* માટે બૂક કરો?\n*Today* અથવા *DD/MM/YYYY* ફોર્મેટમાં તારીખ લખો",
        "confirmed":      "✅ *એપોઇન્ટમેન્ટ કન્ફર્મ!*\n\n👤 *નામ:* {name}\n🔢 *ઉંમર:* {age}\n📋 *કારણ:* {reason}\n📱 *મોબાઇલ:* {mobile}\n📅 *તારીખ:* {date}\n⏰ *સ્લોટ:* {slot}\n\n*જીવનદીપ ક્લિનિક* પસંદ કરવા બદલ આભાર! 🏥",
        "no_slots":       "😔 આ તારીખ માટે કોઈ સ્લોટ ઉપલબ્ધ નથી. અન્ય તારીખ પસંદ કરો:",
        "past_date":      "⚠️ ભૂતકાળની તારીખ માટે બૂક કરી શકાતું નથી. આજ અથવા ભવિષ્યની તારીખ પસંદ કરો:",
        "invalid_date":   "⚠️ અમાન્ય ફોર્મેટ. *Today* અથવા *DD/MM/YYYY* ફોર્મેટ વાપરો (દા.ત. 25/12/2026):",
        "invalid_age":    "⚠️ અમાન્ય ઉંમર. 1 થી 120 વચ્ચેનો નંબર દાખલ કરો:",
        "invalid_mobile": "⚠️ અમાન્ય મોબાઇલ નંબર. બરાબર 10 અંક દાખલ કરો (દા.ત. 9876543210):",
        "invalid_name":   "⚠️ નામ ખૂબ ટૂંકું છે. ઓછામાં ઓછા 2 અક્ષરોનું પૂરું નામ દાખલ કરો:",
        "closed_sunday":  "⚠️ રવિવારે બપોરે 1 વાગ્યા પછી ક્લિનિક બંધ છે. અન્ય તારીખ પસંદ કરો:",
        "reply_yes":      "એપોઇન્ટમેન્ટ માટે *YES* જવાબ આપો. 😊",
        "cancelled":      "❌ બૂકિંગ રદ. ફરી બૂક કરવા /start મોકલો.",
        "no_booking":     "ℹ️ તમારી કોઈ સક્રિય એપોઇન્ટમેન્ট નથી.",
        "your_appt":      "📋 *તમારી એપોઇન્ટમેન્ટ*\n\n👤 {name}\n📅 {date}\n⏰ {slot}\n📋 {reason}",
        "cancel_confirm": "શું તમે ખરેખર {date} ના {slot} ની એપોઇન્ટમેન્ટ રદ કરવા માંગો છો?",
        "cancel_done":    "✅ તમારી એપોઇન્ટમેન્ટ સફળતાપૂર્વક રદ કરવામાં આવી.",
        "reminder_1day":  "⏰ *એપોઇન્ટમેન્ટ રિમાઇન્ડર*\n\nહેલો {name}! જીવનદીપ ક્લિનિકમાં *આવતીકાલ* ({date}) *{slot}* વાગ્યે એપોઇન્ટમેન્ટ છે.\n\n10 મિનિટ વહેલા આવો. 🏥",
        "reminder_1hr":   "⏰ *એપોઇન્ટમેન્ટ રિમાઇન્ડર*\n\nહેલો {name}! *1 કલાક* પછી *{slot}* વાગ્યે એપોઇન્ટમેન્ટ છે.\n\nતૈયાર થઈ જાઓ! 🏥",
    }
}

def S(user_id, key, **kwargs):
    """Get string for user's language."""
    lang = user_languages.get(user_id, "en")
    text = STRINGS[lang].get(key, STRINGS["en"][key])
    return text.format(**kwargs) if kwargs else text

user_languages = {}  # user_id -> "en" or "gu"

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════
#  EXCEL HELPERS
# ══════════════════════════════════════════════

HEADERS = ["User ID", "Name", "Age", "Reason", "Mobile",
           "Date", "Slot Time", "Booking Timestamp", "Status"]

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appointments"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)


def save_appointment(user_id: int, data: dict):
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        user_id,
        data["name"],
        data["age"],
        data["reason"],
        data["mobile"],
        data["date"].strftime("%d/%m/%Y"),
        data["slot_time"],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "ACTIVE",
    ])
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        raise PermissionError("Close appointments.xlsx first.")


def get_row_status(row) -> str:
    """Safely get status from row regardless of old/new format."""
    # New format has 9 cols: User ID, Name, Age, Reason, Mobile, Date, Slot, Timestamp, Status
    # Old format has 7 cols: Name, Age, Reason, Mobile, Date, Slot, Timestamp
    if len(row) >= 9 and row[8] is not None:
        return str(row[8])
    return "ACTIVE"  # Old rows have no status column, treat as ACTIVE


def get_date_col(row) -> int:
    """Returns correct date column index based on row length."""
    return 5 if len(row) >= 9 else 4


def get_slot_col(row) -> int:
    """Returns correct slot column index based on row length."""
    return 6 if len(row) >= 9 else 5


def get_booked_slots(desired_date: date_type) -> list:
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    booked = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if get_row_status(row) == "CANCELLED":
                continue
            date_col = get_date_col(row)
            slot_col = get_slot_col(row)
            booked_date = datetime.strptime(row[date_col], "%d/%m/%Y").date()
            if booked_date == desired_date:
                booked.append(row[slot_col])
        except (ValueError, TypeError, IndexError):
            continue
    return booked


def get_user_appointment(user_id: int):
    """Returns the latest ACTIVE appointment row for a user, or None."""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    result = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if len(row) < 9:
                continue  # Old format rows don't have user_id
            if row[0] == user_id and get_row_status(row) == "ACTIVE":
                appt_date = datetime.strptime(row[5], "%d/%m/%Y").date()
                if appt_date >= datetime.now().date():
                    result = row
        except (ValueError, TypeError, IndexError):
            continue
    return result


def cancel_user_appointment(user_id: int) -> bool:
    """Marks user's latest active future appointment as CANCELLED."""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    cancelled = False
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        if len(vals) < 9:
            continue
        if vals[0] == user_id and vals[8] == "ACTIVE":
            try:
                appt_date = datetime.strptime(vals[5], "%d/%m/%Y").date()
                if appt_date >= datetime.now().date():
                    row[8].value = "CANCELLED"
                    cancelled = True
            except (ValueError, TypeError):
                continue
    if cancelled:
        wb.save(EXCEL_FILE)
    return cancelled


def get_appointments_for_date(target_date: date_type) -> list:
    """Returns all active appointments for a given date, sorted by slot."""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    appts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if get_row_status(row) == "CANCELLED":
                continue
            date_col = get_date_col(row)
            appt_date = datetime.strptime(row[date_col], "%d/%m/%Y").date()
            if appt_date == target_date:
                appts.append(row)
        except (ValueError, TypeError, IndexError):
            continue
    return appts


def get_all_future_active_appointments() -> list:
    """Returns all future active appointments for reminder checks."""
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    appts = []
    today = datetime.now().date()
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if get_row_status(row) == "CANCELLED":
                continue
            date_col = get_date_col(row)
            appt_date = datetime.strptime(row[date_col], "%d/%m/%Y").date()
            if appt_date >= today:
                appts.append(row)
        except (ValueError, TypeError, IndexError):
            continue
    return appts


# ══════════════════════════════════════════════
#  SLOT LOGIC
# ══════════════════════════════════════════════

def get_available_slot(desired_date: date_type):
    WORK_START  = time(9, 0)
    WORK_END    = time(22, 0)
    SUNDAY_END  = time(13, 0)
    LUNCH_START = time(13, 0)
    LUNCH_END   = time(16, 0)
    SLOT_MINS   = timedelta(minutes=30)

    is_sunday     = desired_date.weekday() == 6
    effective_end = SUNDAY_END if is_sunday else WORK_END

    booked_slots = get_booked_slots(desired_date)
    slot_start   = datetime.combine(desired_date, WORK_START)
    end_dt       = datetime.combine(desired_date, effective_end)

    while slot_start < end_dt:
        slot_end = slot_start + SLOT_MINS
        if slot_end > end_dt:
            break
        if not is_sunday and LUNCH_START <= slot_start.time() < LUNCH_END:
            slot_start = datetime.combine(desired_date, LUNCH_END)
            continue
        slot_str = f"{slot_start.strftime('%I:%M %p')} - {slot_end.strftime('%I:%M %p')}"
        if slot_str not in booked_slots:
            return slot_str
        slot_start += SLOT_MINS
    return None


def is_clinic_closed(desired_date: date_type):
    now = datetime.now()
    if desired_date.weekday() == 6:
        if desired_date == now.date() and now.time() >= time(13, 0):
            return True
    return False


# ══════════════════════════════════════════════
#  VALIDATION HELPERS
# ══════════════════════════════════════════════

def validate_age(text: str):
    """Returns (age_int, error_msg). error_msg is None if valid."""
    text = text.strip()
    if not text.isdigit():
        return None, "not_digit"
    age = int(text)
    if age < 1 or age > 120:
        return None, "out_of_range"
    return age, None


def validate_mobile(text: str):
    """
    Valid Indian mobile: exactly 10 digits, starts with 6-9.
    Also accepts +91 prefix.
    Returns (cleaned_number, error_msg).
    """
    text = text.strip()
    # Remove +91 or 91 prefix
    if text.startswith("+91"):
        text = text[3:]
    elif text.startswith("91") and len(text) == 12:
        text = text[2:]
    text = text.replace(" ", "").replace("-", "")
    if not text.isdigit():
        return None, "not_digit"
    if len(text) != 10:
        return None, "not_10_digits"
    if text[0] not in "6789":
        return None, "invalid_start"
    return text, None


def validate_date(text: str):
    """Returns (date_obj, error_msg)."""
    text = text.strip()
    if text.lower() == "today":
        return datetime.now().date(), None
    try:
        d = datetime.strptime(text, "%d/%m/%Y").date()
        return d, None
    except ValueError:
        return None, "invalid_format"


# ══════════════════════════════════════════════
#  LANGUAGE SELECTION
# ══════════════════════════════════════════════

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    keyboard = [[
        InlineKeyboardButton("🇬🇧 English", callback_data="lang_en"),
        InlineKeyboardButton("🇮🇳 ગુજરાતી", callback_data="lang_gu"),
    ]]
    await update.message.reply_text(
        STRINGS["en"]["choose_lang"],
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return CHOOSE_LANG


async def choose_language(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    lang = "en" if query.data == "lang_en" else "gu"
    user_id = query.from_user.id
    user_languages[user_id] = lang

    keyboard = [["YES"]]
    await query.edit_message_text(
        S(user_id, "welcome"),
        parse_mode="Markdown",
    )
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text=S(user_id, "welcome"),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True),
    )
    return WAITING_FOR_YES


# ══════════════════════════════════════════════
#  CONVERSATION HANDLERS
# ══════════════════════════════════════════════

async def waiting_for_yes(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    if update.message.text.strip().upper() == "YES":
        await update.message.reply_text(
            S(user_id, "ask_name"),
            parse_mode="Markdown",
            reply_markup=ReplyKeyboardRemove(),
        )
        return ASK_NAME
    keyboard = [["YES"]]
    await update.message.reply_text(
        S(user_id, "reply_yes"),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True),
    )
    return WAITING_FOR_YES


async def ask_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    name = update.message.text.strip()
    # Validation: at least 2 chars, only letters and spaces
    if len(name) < 2 or not all(c.isalpha() or c.isspace() for c in name):
        await update.message.reply_text(S(user_id, "invalid_name"))
        return ASK_NAME
    context.user_data["name"] = name
    await update.message.reply_text(S(user_id, "ask_age"), parse_mode="Markdown")
    return ASK_AGE


async def ask_age(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    age, err = validate_age(update.message.text)
    if err:
        await update.message.reply_text(S(user_id, "invalid_age"))
        return ASK_AGE
    context.user_data["age"] = age
    await update.message.reply_text(S(user_id, "ask_reason"), parse_mode="Markdown")
    return ASK_REASON


async def ask_reason(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    reason = update.message.text.strip()
    if len(reason) < 3:
        await update.message.reply_text("⚠️ Please describe your reason (at least 3 characters):")
        return ASK_REASON
    context.user_data["reason"] = reason
    await update.message.reply_text(S(user_id, "ask_mobile"), parse_mode="Markdown")
    return ASK_MOBILE


async def ask_mobile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    mobile, err = validate_mobile(update.message.text)
    if err:
        await update.message.reply_text(S(user_id, "invalid_mobile"))
        return ASK_MOBILE
    context.user_data["mobile"] = mobile
    keyboard = [["Today"]]
    await update.message.reply_text(
        S(user_id, "ask_date"),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True),
    )
    return ASK_DATE


async def ask_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    text  = update.message.text.strip()
    today = datetime.now().date()

    desired_date, err = validate_date(text)
    if err:
        await update.message.reply_text(
            S(user_id, "invalid_date"), parse_mode="Markdown"
        )
        return ASK_DATE

    if desired_date < today:
        await update.message.reply_text(
            S(user_id, "past_date"), reply_markup=ReplyKeyboardRemove()
        )
        return ASK_DATE

    if is_clinic_closed(desired_date):
        await update.message.reply_text(
            S(user_id, "closed_sunday"), reply_markup=ReplyKeyboardRemove()
        )
        return ASK_DATE

    available_slot = get_available_slot(desired_date)
    if not available_slot:
        # Suggest tomorrow automatically
        tomorrow = desired_date + timedelta(days=1)
        tomorrow_slot = get_available_slot(tomorrow)
        if tomorrow_slot:
            keyboard = [[tomorrow.strftime("%d/%m/%Y"), "Other Date"]]
            await update.message.reply_text(
                f"😔 No slots available for {desired_date.strftime('%d/%m/%Y')}.

"
                f"✅ *Tomorrow ({tomorrow.strftime('%d/%m/%Y')})* has slots available!
"
                f"Reply with tomorrow's date or enter another date (DD/MM/YYYY):",
                parse_mode="Markdown",
                reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True),
            )
        else:
            await update.message.reply_text(
                S(user_id, "no_slots"), reply_markup=ReplyKeyboardRemove()
            )
        return ASK_DATE

    context.user_data["date"]      = desired_date
    context.user_data["slot_time"] = available_slot

    try:
        save_appointment(user_id, context.user_data)
    except PermissionError:
        await update.message.reply_text(
            "⚠️ Could not save appointment. Please try again.",
            reply_markup=ReplyKeyboardRemove(),
        )
        return ConversationHandler.END

    await update.message.reply_text(
        S(user_id, "confirmed",
          name=context.user_data["name"],
          age=context.user_data["age"],
          reason=context.user_data["reason"],
          mobile=context.user_data["mobile"],
          date=desired_date.strftime("%d/%m/%Y"),
          slot=available_slot),
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )

    # Notify owner
    try:
        await context.bot.send_message(
            chat_id=OWNER_ID,
            text=(
                f"🔔 *New Appointment*\n\n"
                f"👤 {context.user_data['name']} | Age {context.user_data['age']}\n"
                f"📱 {context.user_data['mobile']}\n"
                f"📋 {context.user_data['reason']}\n"
                f"📅 {desired_date.strftime('%d/%m/%Y')} @ {available_slot}"
            ),
            parse_mode="Markdown",
        )
    except Exception:
        pass

    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    context.user_data.clear()
    await update.message.reply_text(
        S(user_id, "cancelled"), reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END


# ══════════════════════════════════════════════
#  PATIENT SELF-SERVICE
# ══════════════════════════════════════════════

async def my_appointment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    appt = get_user_appointment(user_id)
    if not appt:
        await update.message.reply_text(S(user_id, "no_booking"))
        return
    await update.message.reply_text(
        S(user_id, "your_appt",
          name=appt[1], date=appt[5], slot=appt[6], reason=appt[3]),
        parse_mode="Markdown",
    )


async def cancel_appointment_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    appt = get_user_appointment(user_id)
    if not appt:
        await update.message.reply_text(S(user_id, "no_booking"))
        return
    keyboard = [[
        InlineKeyboardButton("✅ Yes, Cancel", callback_data="confirm_cancel"),
        InlineKeyboardButton("❌ No, Keep it", callback_data="keep_appt"),
    ]]
    await update.message.reply_text(
        S(user_id, "cancel_confirm", date=appt[5], slot=appt[6]),
        reply_markup=InlineKeyboardMarkup(keyboard),
    )


async def handle_cancel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if query.data == "confirm_cancel":
        success = cancel_user_appointment(user_id)
        if success:
            await query.edit_message_text(S(user_id, "cancel_done"))
            try:
                appt = get_user_appointment(user_id)
                name = appt[1] if appt else "Patient"
                await context.bot.send_message(
                    chat_id=OWNER_ID,
                    text=f"🚫 *Appointment Cancelled*\n\n👤 Patient cancelled their appointment.",
                    parse_mode="Markdown",
                )
            except Exception:
                pass
        else:
            await query.edit_message_text("⚠️ Could not cancel. Please try again.")
    else:
        await query.edit_message_text("✅ Your appointment is kept. See you soon! 🏥")


# ══════════════════════════════════════════════
#  OWNER COMMANDS
# ══════════════════════════════════════════════

def format_schedule(appts: list, title: str) -> str:
    if not appts:
        return f"📅 *{title}*\n\nNo appointments."
    # Sort by slot time
    appts_sorted = sorted(appts, key=lambda r: r[6])
    lines = [f"📅 *{title}* — {len(appts_sorted)} appointment(s)\n"]
    for i, row in enumerate(appts_sorted, 1):
        lines.append(f"{i}. ⏰ {row[6]}\n   👤 {row[1]} | Age {row[2]}\n   📱 {row[4]} | 📋 {row[3]}\n")
    return "\n".join(lines)


async def today_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID:
        return
    appts = get_appointments_for_date(datetime.now().date())
    msg   = format_schedule(appts, f"Today's Schedule ({datetime.now().strftime('%d/%m/%Y')})")
    await update.message.reply_text(msg, parse_mode="Markdown")


async def tomorrow_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID:
        return
    tomorrow = datetime.now().date() + timedelta(days=1)
    appts    = get_appointments_for_date(tomorrow)
    msg      = format_schedule(appts, f"Tomorrow's Schedule ({tomorrow.strftime('%d/%m/%Y')})")
    await update.message.reply_text(msg, parse_mode="Markdown")


async def send_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID:
        return
    if os.path.exists(EXCEL_FILE):
        await update.message.reply_document(
            document=open(EXCEL_FILE, "rb"),
            filename=EXCEL_FILE,
            caption="📊 All appointments.",
        )
    else:
        await update.message.reply_text("ℹ️ No appointments yet.")


async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID:
        return
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    today_str = datetime.now().strftime("%d/%m/%Y")
    total, today_count, cancelled_count = 0, 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[8] == "CANCELLED":
            cancelled_count += 1
            continue
        total += 1
        if row[5] == today_str:
            today_count += 1
    await update.message.reply_text(
        f"📊 *Appointment Stats*\n\n"
        f"✅ Total active  : {total}\n"
        f"❌ Cancelled     : {cancelled_count}\n"
        f"📅 Today ({today_str}): {today_count}",
        parse_mode="Markdown",
    )


async def search_patient(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID:
        return
    if not context.args:
        await update.message.reply_text("Usage: /search <name>\nExample: /search Rahul")
        return
    query_name = " ".join(context.args).lower()
    initialize_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and query_name in str(row[1]).lower():
            results.append(row)
    if not results:
        await update.message.reply_text(f"No results for '{query_name}'.")
        return
    lines = [f"🔍 *Search: {query_name}* — {len(results)} found\n"]
    for r in results:
        status = "✅" if r[8] == "ACTIVE" else "❌"
        lines.append(f"{status} {r[1]} | {r[5]} @ {r[6]}\n   📱 {r[4]} | 📋 {r[3]}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ══════════════════════════════════════════════
#  SCHEDULED JOBS
# ══════════════════════════════════════════════

async def send_daily_summary(bot):
    """Sends daily summary to owner at 10 PM."""
    try:
        today    = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        today_appts    = get_appointments_for_date(today)
        tomorrow_appts = get_appointments_for_date(tomorrow)

        msg = (
            f"🌙 *Daily Summary — {today.strftime('%d/%m/%Y')}*\n\n"
            f"Today's patients    : {len(today_appts)}\n"
            f"Tomorrow's patients : {len(tomorrow_appts)}\n"
        )
        if tomorrow_appts:
            first = sorted(tomorrow_appts, key=lambda r: r[6])[0]
            msg += f"\nFirst slot tomorrow : {first[6]} — {first[1]}"

        await bot.send_message(chat_id=OWNER_ID, text=msg, parse_mode="Markdown")
    except Exception as e:
        logger.error(f"Daily summary error: {e}")


async def send_reminders(bot):
    """Checks every 30 min and sends 1-day and 1-hour reminders."""
    try:
        now       = datetime.now()
        tomorrow  = (now + timedelta(days=1)).date()
        appts     = get_all_future_active_appointments()

        for appt in appts:
            user_id    = appt[0]
            name       = appt[1]
            appt_date  = datetime.strptime(appt[5], "%d/%m/%Y").date()
            slot_start = appt[6].split(" - ")[0].strip()  # e.g. "10:30 AM"

            try:
                appt_dt = datetime.combine(
                    appt_date,
                    datetime.strptime(slot_start, "%I:%M %p").time()
                )
            except ValueError:
                continue

            diff_hours = (appt_dt - now).total_seconds() / 3600

            # 1-day reminder: between 23h and 25h before
            if 23 <= diff_hours <= 25:
                lang = user_languages.get(user_id, "en")
                msg  = STRINGS[lang]["reminder_1day"].format(
                    name=name, date=appt[5], slot=appt[6]
                )
                try:
                    await bot.send_message(chat_id=user_id, text=msg, parse_mode="Markdown")
                except Exception:
                    pass

            # 1-hour reminder: between 0.9h and 1.1h before
            elif 0.9 <= diff_hours <= 1.1:
                lang = user_languages.get(user_id, "en")
                msg  = STRINGS[lang]["reminder_1hr"].format(
                    name=name, slot=appt[6]
                )
                try:
                    await bot.send_message(chat_id=user_id, text=msg, parse_mode="Markdown")
                except Exception:
                    pass

    except Exception as e:
        logger.error(f"Reminder error: {e}")


# ══════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════

def main():
    if not BOT_TOKEN:
        raise ValueError("BOT_TOKEN environment variable not set!")
    if not OWNER_ID:
        raise ValueError("OWNER_ID environment variable not set!")

    initialize_excel()

    app = Application.builder().token(BOT_TOKEN).build()

    # Conversation
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            CHOOSE_LANG     : [CallbackQueryHandler(choose_language, pattern="^lang_")],
            WAITING_FOR_YES : [MessageHandler(filters.TEXT & ~filters.COMMAND, waiting_for_yes)],
            ASK_NAME        : [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_name)],
            ASK_AGE         : [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_age)],
            ASK_REASON      : [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_reason)],
            ASK_MOBILE      : [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_mobile)],
            ASK_DATE        : [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_date)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )

    app.add_handler(conv_handler)

    # Patient commands
    app.add_handler(CommandHandler("my_appointment",     my_appointment))
    app.add_handler(CommandHandler("cancel_appointment", cancel_appointment_cmd))
    app.add_handler(CallbackQueryHandler(handle_cancel_callback, pattern="^(confirm_cancel|keep_appt)$"))

    # Owner commands
    app.add_handler(CommandHandler("today",    today_schedule))
    app.add_handler(CommandHandler("tomorrow", tomorrow_schedule))
    app.add_handler(CommandHandler("send",     send_excel))
    app.add_handler(CommandHandler("share",    send_excel))
    app.add_handler(CommandHandler("stats",    stats))
    app.add_handler(CommandHandler("search",   search_patient))

    # Scheduler
    scheduler = AsyncIOScheduler()
    # Daily summary at 10 PM
    scheduler.add_job(
        send_daily_summary, "cron", hour=22, minute=0,
        args=[app.bot]
    )
    # Reminders check every 30 minutes
    scheduler.add_job(
        send_reminders, "interval", minutes=30,
        args=[app.bot]
    )
    scheduler.start()

    print("🏥 Jivandeep Clinic Bot is starting...")
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)


if __name__ == "__main__":
    main()
